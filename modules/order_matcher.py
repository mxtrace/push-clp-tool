"""
order_matcher.py — 模块2：订单类型匹配（MSPP + SMP）
从本地 Loading 表读取 AL0 列表，标记 MSPP 和 SMP，输出目标订单列表。
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Optional

import openpyxl


# PRD v1.1：OC Task 未发送状态值
NOT_SENT_STATUSES: frozenset[str] = frozenset({"PENDING", "CLOSE_TO_SLA", "MISSED_SLA", "NO_SLA"})


@dataclass
class OrderRecord:
    """Loading 表中的一条订单记录，附带 MSPP/SMP 标记。"""
    al0:             str
    pol:             str
    pod:             str
    flipped_fc:      str
    si_sent:         bool           # True = Loading 表 SI_SENT 非空
    customs_seller:  bool           # True = Loading 表 报关资料-CHB 非空（保留用于日志）
    shipper_id:      str
    is_mspp:         bool = False
    is_smp:          bool = False
    ops_login:       str  = ""
    sales_login:     str  = ""
    # PRD v1.1 新增：OC Task 状态（由 fetch_pending_tasks 填充）
    si_oc_status:       str = ""    # SEND_SI_TO_LSP_LCL 最新状态，"" 表示无任务
    customs_oc_status:  str = ""    # SELLER_SEND_EXPORT_DOC_LCL 最新状态

    @property
    def is_target(self) -> bool:
        """目标订单 = MSPP AND SMP。"""
        return self.is_mspp and self.is_smp

    @property
    def si_not_sent(self) -> bool:
        """PRD v1.1 双渠道：Loading 表 SI_SENT 为空 AND OC SI Task 处于未发送状态。"""
        return not self.si_sent and self.si_oc_status in NOT_SENT_STATUSES

    @property
    def customs_not_sent(self) -> bool:
        """PRD v1.1：OC 报关 Task 处于未发送状态。"""
        return self.customs_oc_status in NOT_SENT_STATUSES

    @property
    def needs_push(self) -> bool:
        """需要催交 = 是目标订单 AND（SI 未发 OR 报关未提交）。"""
        return self.is_target and (self.si_not_sent or self.customs_not_sent)


def load_loading_table(
    file_path: str,
    cols: dict[str, str],
    sheet_name: Optional[str] = None,
    header_row: int = 1,
) -> list[OrderRecord]:
    """
    读取本地 BC Loading 表（Excel），返回 OrderRecord 列表。

    Args:
        file_path:  本地 Excel 文件路径
        cols:       列名映射 dict，key = 标准名，value = Excel 实际列标题
                    必填 key: al0, pol, flipped_fc, si_sent, customs_seller, shipper_id
        sheet_name: 指定 sheet 名（默认读第一个 sheet）
        header_row: 表头所在行号（1-based，默认 1）
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = header_row - 1   # 转为 0-based
    if header_idx >= len(rows):
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    col_idx = {v: i for i, v in enumerate(header)}

    def _idx(col_key: str) -> Optional[int]:
        col_name = cols.get(col_key, col_key)
        return col_idx.get(col_name)

    records = []
    for row in rows[header_idx + 1:]:
        al0_val = _cell(row, _idx("al0"))
        if not al0_val:
            continue          # 跳过空行
        records.append(OrderRecord(
            al0=al0_val,
            pol=_cell(row, _idx("pol")),
            pod=_cell(row, _idx("pod")),
            flipped_fc=_cell(row, _idx("flipped_fc")),
            si_sent=_is_filled(row, _idx("si_sent")),
            customs_seller=_is_filled(row, _idx("customs_seller")),
            shipper_id=_cell(row, _idx("shipper_id")),
        ))
    wb.close()
    return records


def load_mspp_shipper_list(
    file_path: str,
    sheet_name: str = "Summary",
    shipper_id_col: str = "shipper_id",
) -> set[str]:
    """
    读取 MSPP Shipper List Excel，返回 shipper_id 集合（用于快速 membership check）。
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return set()

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    try:
        idx = header.index(shipper_id_col)
    except ValueError:
        wb.close()
        raise ValueError(f"MSPP Shipper List 中未找到列 '{shipper_id_col}'，实际列: {header}")

    shipper_ids = set()
    for row in rows[1:]:
        val = _cell(row, idx)
        if val:
            shipper_ids.add(val)
    wb.close()
    return shipper_ids


def apply_mspp_flag(
    orders: list[OrderRecord],
    mspp_shipper_ids: set[str],
) -> None:
    """原地更新 OrderRecord.is_mspp 标记（in-place）。"""
    for order in orders:
        order.is_mspp = order.shipper_id in mspp_shipper_ids


def apply_smp_and_contact_flags(
    orders: list[OrderRecord],
    cache_data: dict[str, dict],
) -> None:
    """
    原地更新 is_smp、ops_login、sales_login（in-place）。
    同时从缓存恢复 si_oc_status / customs_oc_status（如有）。
    """
    for order in orders:
        row = cache_data.get(order.al0, {})
        order.is_smp          = bool(row.get("is_smp", False))
        order.ops_login       = str(row.get("ops_login", "")).strip()
        order.sales_login     = str(row.get("sales_login", "")).strip()
        order.si_oc_status    = str(row.get("si_task_status", "")).strip()
        order.customs_oc_status = str(row.get("customs_task_status", "")).strip()


def apply_task_statuses(
    orders: list[OrderRecord],
    task_data: list[dict],
) -> None:
    """
    原地更新 si_oc_status / customs_oc_status（PRD v1.1 双渠道核验）。

    Args:
        orders:    订单列表
        task_data: fetch_pending_tasks() 的返回值
                   [{al0, si_creation_date, si_task_status, customs_task_status}]
    """
    task_map = {t["al0"]: t for t in task_data}
    for order in orders:
        info = task_map.get(order.al0, {})
        order.si_oc_status      = str(info.get("si_task_status",      "")).strip()
        order.customs_oc_status = str(info.get("customs_task_status", "")).strip()


# ------------------------------------------------------------------ #
# 内部工具
# ------------------------------------------------------------------ #

def _cell(row: tuple, idx: Optional[int]) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _is_filled(row: tuple, idx: Optional[int]) -> bool:
    """单元格非空视为"已填写"（SI 已发送 / 报关资料已提交）。"""
    if idx is None or idx >= len(row):
        return False
    return row[idx] is not None and str(row[idx]).strip() != ""
