"""
status_writer.py — 生成 push_status.xlsx 状态诊断表
Sheet1: 筛选漏斗（全部目标订单的各步骤状态）
Sheet2: 推送清单（最终进入 clp_items 的订单）
"""
from __future__ import annotations

from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# 颜色
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True)


def _determine_final_status(trace, pushed_al0s: set) -> tuple:
    """返回 (最终状态, 排除原因)。"""
    if trace.al0 in pushed_al0s:
        return "pushed", ""
    if trace.step_b != "PASS":
        return "SKIP_B", trace.step_b_detail or trace.step_b
    if trace.step_c != "PASS":
        return "SKIP_C", trace.step_c_detail or trace.step_c

    if trace.step_e != "PASS":
        return "SKIP_E", trace.step_e_detail or trace.step_e
    return trace.result, ""


def write_status(
    orders,
    traces,
    clp_items,
    output_path: str,
    run_time: datetime,
):
    """生成 push_status.xlsx 双 Sheet 诊断表。"""
    run_time_str = run_time.strftime("%Y-%m-%d %H:%M:%S")
    pushed_al0s = {item.al0 for item in clp_items}

    wb = openpyxl.Workbook()

    # ── Sheet1: 筛选漏斗 ──
    ws1 = wb.active
    ws1.title = "筛选漏斗"
    headers1 = [
        "AL0", "POL", "POD", "Shipper_ID", "is_MSPP", "is_SMP",
        "ETD", "Vessel", "Voyage", "SI_Cutoff", "CLP_Cutoff", "Service",
        "step_B", "step_B_detail",
        "step_C", "step_C_detail",
        "step_D", "step_D_detail",
        "step_E", "step_E_detail",
        "最终状态", "排除原因", "run_time",
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = HEADER_FONT

    # 建立 trace 索引
    trace_map = {t.al0: t for t in traces}

    for order in orders:
        trace = trace_map.get(order.al0)
        if trace:
            final_status, reason = _determine_final_status(trace, pushed_al0s)
            row = [
                order.al0, order.pol, order.pod, order.shipper_id,
                order.is_mspp, order.is_smp,
                trace.matched_etd, trace.matched_vessel, trace.matched_voyage,
                trace.matched_si_cutoff, trace.matched_clp_cutoff, trace.matched_service,
                trace.step_b, trace.step_b_detail,
                trace.step_c, trace.step_c_detail,
                trace.step_d, trace.step_d_detail,
                trace.step_e, trace.step_e_detail,
                final_status, reason, run_time_str,
            ]
        else:
            # 未进入 Task2 的订单（非目标）
            final_status = "NOT_TARGET"
            reason = "非MSPP或非SMP"
            if not order.is_mspp:
                reason = "非MSPP"
            elif not order.is_smp:
                reason = "非SMP"
            row = [
                order.al0, order.pol, order.pod, order.shipper_id,
                order.is_mspp, order.is_smp,
                "", "", "", "", "", "",
                "", "", "", "", "", "", "", "",
                final_status, reason, run_time_str,
            ]

        ws1.append(row)
        row_idx = ws1.max_row

        # 颜色规则
        if final_status == "pushed":
            for cell in ws1[row_idx]:
                cell.fill = GREEN_FILL
        elif "SKIP" in final_status or "FAIL" in final_status:
            for cell in ws1[row_idx]:
                cell.fill = YELLOW_FILL
        elif not order.is_mspp or not order.is_smp:
            for cell in ws1[row_idx]:
                cell.fill = GREY_FILL

    # ── Sheet2: 推送清单 ──
    ws2 = wb.create_sheet("推送清单")
    headers2 = [
        "AL0", "POL", "POD", "ETD", "SI_Cutoff",
        "all_ready_datetime", "push_status", "run_time",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = HEADER_FONT

    for item in clp_items:
        row = [
            item.al0, item.pol, item.pod,
            item.etd_str, item.si_cutoff_str,
            item.all_ready_str, "pushed", run_time_str,
        ]
        ws2.append(row)
        for cell in ws2[ws2.max_row]:
            cell.fill = GREEN_FILL

    # 列宽自适应
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(output_path)
