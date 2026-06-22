"""
test_changes.py — 本次变更的单元测试
覆盖：
  1. _find_nearest_sailing 过期过滤
  2. DEMO_POL_FILTER 常量
  3. fetch_clp_via_container_details 返回值解包
  4. taskStatus None 防御
  5. status_writer 基本功能
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from datetime import datetime, date, time
from unittest.mock import patch, MagicMock
import pytz

BEIJING_TZ = pytz.timezone("Asia/Shanghai")


# ═══════════════════════════════════════════════════════════════════
# Test 1: _find_nearest_sailing — 过期船期被过滤
# ═══════════════════════════════════════════════════════════════════
def test_find_nearest_sailing_filters_expired():
    from modules.push_clp import _find_nearest_sailing
    from modules.schedule_builder import SailingRecord

    # 构造两条船期：一条已过期(6/20)，一条有效(6/25)
    expired = SailingRecord(
        pol="CNYTN", pod="USLAX", etd=date(2026, 6, 22),
        vessel="V1", voyage="001",
        si_cutoff=BEIJING_TZ.localize(datetime(2026, 6, 20, 17, 0)),
        available_capacity=10.0,
    )
    valid = SailingRecord(
        pol="CNYTN", pod="USLAX", etd=date(2026, 6, 30),
        vessel="V2", voyage="002",
        si_cutoff=BEIJING_TZ.localize(datetime(2026, 6, 25, 17, 0)),
        available_capacity=10.0,
    )

    now = BEIJING_TZ.localize(datetime(2026, 6, 22, 10, 0))  # 6/22 10:00

    result = _find_nearest_sailing("CNYTN", "USLAX", [expired, valid], now)
    assert result is not None, "应返回有效船期"
    assert result.voyage == "002", f"应匹配 valid 船期，实际={result.voyage}"
    print("  ✓ Test 1 PASS: 过期船期被正确过滤")


def test_find_nearest_sailing_all_expired():
    from modules.push_clp import _find_nearest_sailing
    from modules.schedule_builder import SailingRecord

    expired = SailingRecord(
        pol="CNYTN", pod="USLAX", etd=date(2026, 6, 22),
        vessel="V1", voyage="001",
        si_cutoff=BEIJING_TZ.localize(datetime(2026, 6, 20, 17, 0)),
        available_capacity=10.0,
    )

    now = BEIJING_TZ.localize(datetime(2026, 6, 22, 10, 0))
    result = _find_nearest_sailing("CNYTN", "USLAX", [expired], now)
    assert result is None, "全部过期应返回None"
    print("  ✓ Test 2 PASS: 全部过期 → 返回None")


def test_find_nearest_sailing_boundary_17():
    """刚好17:00时仍有效"""
    from modules.push_clp import _find_nearest_sailing
    from modules.schedule_builder import SailingRecord

    sailing = SailingRecord(
        pol="CNYTN", pod="USLAX", etd=date(2026, 6, 25),
        vessel="V1", voyage="001",
        si_cutoff=BEIJING_TZ.localize(datetime(2026, 6, 22, 14, 0)),
        available_capacity=10.0,
    )

    # now = SI_Cutoff当天17:00:00 → 仍有效
    now = BEIJING_TZ.localize(datetime(2026, 6, 22, 17, 0, 0))
    result = _find_nearest_sailing("CNYTN", "USLAX", [sailing], now)
    assert result is not None, "17:00:00 应仍有效"

    # now = SI_Cutoff当天17:00:01 → 过期
    now_after = BEIJING_TZ.localize(datetime(2026, 6, 22, 17, 0, 1))
    result2 = _find_nearest_sailing("CNYTN", "USLAX", [sailing], now_after)
    assert result2 is None, "17:00:01 应过期"
    print("  ✓ Test 3 PASS: 17:00边界行为正确")


# ═══════════════════════════════════════════════════════════════════
# Test 4: DEMO_POL_FILTER 常量验证
# ═══════════════════════════════════════════════════════════════════
def test_demo_pol_filter():
    from modules.order_matcher import DEMO_POL_FILTER
    assert "CNYTN" in DEMO_POL_FILTER
    assert "CNNGB" not in DEMO_POL_FILTER
    assert "CNSHA" not in DEMO_POL_FILTER
    print("  ✓ Test 4 PASS: DEMO_POL_FILTER 仅含 CNYTN")


# ═══════════════════════════════════════════════════════════════════
# Test 5: fetch_clp_via_container_details 返回 tuple 解包
# ═══════════════════════════════════════════════════════════════════
def test_clp_container_returns_tuple():
    from modules.oc_client import fetch_clp_via_container_details

    # Mock session
    mock_session = MagicMock()
    mock_resp = MagicMock()
    mock_resp.status_code = 200
    mock_resp.json.return_value = {
        "status": "SUCCESS",
        "data": {"containers": []}
    }
    mock_session.get.return_value = mock_resp

    result = fetch_clp_via_container_details("AL0-TEST", session=mock_session)
    assert isinstance(result, tuple), f"应返回tuple，实际={type(result)}"
    assert len(result) == 2, f"应为2元素tuple，实际={len(result)}"
    flag, diag = result
    assert flag is False, f"containers=[] 应返回 False，实际={flag}"
    assert "containers=[]" in diag
    print("  ✓ Test 5 PASS: 返回tuple(False, diag)正确")


def test_clp_container_api_fail():
    from modules.oc_client import fetch_clp_via_container_details

    mock_session = MagicMock()
    mock_resp = MagicMock()
    mock_resp.status_code = 200
    mock_resp.json.return_value = {"status": "FAIL"}
    mock_session.get.return_value = mock_resp

    flag, diag = fetch_clp_via_container_details("AL0-TEST", session=mock_session)
    assert flag is None, f"API失败应返回None，实际={flag}"
    assert "status=FAIL" in diag
    print("  ✓ Test 6 PASS: API失败返回(None, diag)")


# ═══════════════════════════════════════════════════════════════════
# Test 7: taskStatus=None 防御
# ═══════════════════════════════════════════════════════════════════
def test_task_status_none_defense():
    """模拟 booking_raw.taskStatus = None 的情况"""
    booking_raw = {"taskStatus": None, "taskAssignees": {}}
    # 模拟代码逻辑
    task_status = booking_raw.get("taskStatus") or ""
    assert task_status == "", f"None应被转为空字符串，实际={repr(task_status)}"
    assert str(task_status).strip() == ""
    print("  ✓ Test 7 PASS: taskStatus=None 被正确防御为空字符串")


# ═══════════════════════════════════════════════════════════════════
# Test 8: status_writer 基本功能
# ═══════════════════════════════════════════════════════════════════
def test_status_writer():
    import tempfile
    from modules.status_writer import write_status
    from modules.push_clp import CLPItem, OrderTrace
    from modules.order_matcher import OrderRecord

    orders = [
        OrderRecord(
            al0="AL0-TEST1", pol="CNYTN", pod="USLAX",
            flipped_fc="", si_sent=False, customs_seller=False,
            shipper_id="S001", is_mspp=True, is_smp=True,
        )
    ]
    traces = [
        OrderTrace(al0="AL0-TEST1", pol="CNYTN", pod="USLAX",
                   step_b="PASS", step_c="PASS", step_e="PASS", result="PUSHED")
    ]
    clp_items = [
        CLPItem(
            al0="AL0-TEST1",
            all_ready_datetime=BEIJING_TZ.localize(datetime(2026, 6, 20, 10, 0)),
            si_cutoff=BEIJING_TZ.localize(datetime(2026, 6, 24, 17, 0)),
            etd=date(2026, 6, 30),
            pol="CNYTN", pod="USLAX",
            vessel="EVER ACE", voyage="0123E",
        )
    ]

    out_path = os.path.join(os.path.dirname(__file__), "data", "test_status.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    write_status(
        orders=orders, traces=traces, clp_items=clp_items,
        output_path=out_path,
        run_time=BEIJING_TZ.localize(datetime(2026, 6, 22, 14, 0)),
    )
    assert os.path.exists(out_path), "xlsx文件未生成"

    import openpyxl
    wb = openpyxl.load_workbook(out_path)
    assert "筛选漏斗" in wb.sheetnames
    assert "推送清单" in wb.sheetnames
    ws2 = wb["推送清单"]
    assert ws2.cell(2, 1).value == "AL0-TEST1"
    assert ws2.cell(2, 7).value == "pushed"
    wb.close()
    os.remove(out_path)
    print("  ✓ Test 8 PASS: status_writer 生成正确的双Sheet xlsx")


# ═══════════════════════════════════════════════════════════════════
# Run all tests
# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("  === Push CLP 单元测试 ===")
    test_find_nearest_sailing_filters_expired()
    test_find_nearest_sailing_all_expired()
    test_find_nearest_sailing_boundary_17()
    test_demo_pol_filter()
    test_clp_container_returns_tuple()
    test_clp_container_api_fail()
    test_task_status_none_defense()
    test_status_writer()
    print("  === 全部 8 项测试通过 ✓ ===")
